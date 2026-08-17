#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generador de descompuestos en formato de hoja de descompuesto (estilo CYPE).

El importador del proyecto (`app/services/importer.py`) NO reconoce un
fichero por su origen, sino por su ESTRUCTURA: busca una fila de encabezados
con «Código, Unidad, Descripción, Rendimiento, Precio unitario, Importe» y a
partir de ahí deduce las columnas. Por tanto podemos escribir nosotros ese
mismo layout con datos 100% propios y la aplicación lo detecta igual de bien,
con todos los rendimientos por hora / m² de mano de obra, materiales y
maquinaria intactos.

Este módulo es la pieza que convierte una partida descrita en YAML/JSON
(datos/descompuestos/*.json) en el .xlsx que la app ingiere.

Uso:
    python3 basedatos_partidas/descompuestos.py            # genera todos
    python3 basedatos_partidas/descompuestos.py PRO-ALB-001
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent
RAIZ = BASE.parent
ORIGEN = BASE / "datos" / "descompuestos"
RECURSOS = BASE / "datos" / "recursos.json"
CLASIFICACION = BASE / "datos" / "clasificacion.json"
SALIDA = BASE / "salida" / "descompuestos"

# Margen por defecto para pasar de coste directo a precio de venta del catálogo.
MARGEN_DEFECTO = 0.30

# Layout de 8 columnas (como DPT020.xlsx): A Código, D Unidad, E Descripción,
# F Rendimiento, G Precio unitario, H Importe.
COL_CODIGO, COL_UNIDAD, COL_DESCRIPCION = 1, 4, 5
COL_RENDIMIENTO, COL_PRECIO, COL_IMPORTE = 6, 7, 8

# Etiquetas de grupo que el clasificador del proyecto entiende
# (_categoria_coste_cype): materiales / equipo y maquinaria / mano de obra /
# costes directos complementarios.
ORDEN_GRUPOS = [
    ("materiales", "Materiales"),
    ("maquinaria", "Equipo y maquinaria"),
    ("mano_obra", "Mano de obra"),
]

F_IMPORTE = "=ROUND(INDIRECT(ADDRESS(ROW()+(0), COLUMN()+(-2), 1))*INDIRECT(ADDRESS(ROW()+(0), COLUMN()+(-1), 1)), 2)"


def _f_subtotal(n: int) -> str:
    partes = ",".join(
        f"INDIRECT(ADDRESS(ROW()+({-i}), COLUMN()+(0), 1))" for i in range(1, n + 1)
    )
    return f"=ROUND(SUM({partes}), 2)"


def _f_base_complementarios(offsets: list[int]) -> str:
    partes = ",".join(
        f"INDIRECT(ADDRESS(ROW()+({o}), COLUMN()+(1), 1))" for o in offsets
    )
    return f"=ROUND(SUM({partes}), 2)"


def _f_total(offsets: list[int]) -> str:
    partes = ",".join(
        f"INDIRECT(ADDRESS(ROW()+({o}), COLUMN()+(0), 1))" for o in offsets
    )
    return f"=ROUND(SUM({partes}), 2)"


F_COMPLEMENTARIOS = "=ROUND(INDIRECT(ADDRESS(ROW()+(0), COLUMN()+(-2), 1))*INDIRECT(ADDRESS(ROW()+(0), COLUMN()+(-1), 1))/100, 2)"


def cargar_clasificacion() -> dict:
    return json.loads(CLASIFICACION.read_text(encoding="utf-8")) if CLASIFICACION.exists() else {}


def ubicar(partida: dict, taxonomia: dict) -> dict:
    """Valida la ruta numérica de tres niveles y devuelve nombres legibles.

    La taxonomía v2 usa capítulo, subcapítulo y apartado. El código visible
    ``CC.SS.AA.NNN`` debe coincidir con los tres segmentos declarados en la
    partida; el antiguo ``CT-CC-SS-NNN`` solo se conserva como trazabilidad.
    """
    caps = taxonomia.get("capitulos", {})
    cap = str(partida.get("capitulo") or "")
    sub = str(partida.get("subcapitulo") or "")
    apartado = str(partida.get("apartado") or "")
    if cap not in caps:
        raise ValueError(f"{partida['codigo']}: capítulo «{cap}» no existe en clasificacion.json")
    subs = caps[cap].get("subcapitulos") or {}
    if sub not in subs:
        raise ValueError(f"{partida['codigo']}: subcapítulo «{sub}» no existe en el capítulo {cap}")
    nodo_sub = subs[sub]
    apartados = nodo_sub.get("apartados") or {}
    if apartado not in apartados:
        raise ValueError(
            f"{partida['codigo']}: apartado «{apartado}» no existe en {cap}.{sub}"
        )
    esperado = f"{cap}.{sub}.{apartado}."
    if not str(partida["codigo"]).startswith(esperado):
        raise ValueError(f"{partida['codigo']}: el código debería empezar por «{esperado}»")
    ambito = partida.get("ambito", "reforma")
    if ambito != taxonomia.get("_ambito", "reforma"):
        raise ValueError(f"{partida['codigo']}: ámbito «{ambito}» no corresponde a esta clasificación")
    return {
        "capitulo_cod": cap,
        "capitulo": caps[cap]["nombre"],
        "subcapitulo_cod": sub,
        "subcapitulo": nodo_sub["nombre"],
        "apartado_cod": apartado,
        "apartado": apartados[apartado],
        "ambito": ambito,
    }


def cargar_recursos() -> dict:
    """Aplana datos/recursos.json en un mapa codigo -> ficha del recurso.

    Los recursos que declaran `composicion` (morteros y concretos elaborados
    en obra) **no llevan precio propio**: se calcula sumando sus componentes.
    Así, subir el cemento recalcula solo el mortero de pega, el de friso y el
    de contrapiso, sin tener que acordarse de tocarlos uno a uno.
    """
    if not RECURSOS.exists():
        return {}
    bruto = json.loads(RECURSOS.read_text(encoding="utf-8"))
    plano: dict[str, dict] = {}
    for grupo in ("materiales", "maquinaria", "mano_obra"):
        for codigo, ficha in (bruto.get(grupo) or {}).items():
            plano[codigo] = {**ficha, "grupo": grupo, "codigo": codigo}

    # Precio de los compuestos, resuelto en cascada y detectando ciclos.
    def precio_de(codigo: str, visitando: tuple[str, ...] = ()) -> float:
        ficha = plano.get(codigo)
        if ficha is None:
            raise ValueError(f"el recurso «{codigo}» no existe en recursos.json")
        composicion = ficha.get("composicion")
        if not composicion:
            return float(ficha["precio"])
        if codigo in visitando:
            ciclo = " -> ".join((*visitando, codigo))
            raise ValueError(f"composición circular de recursos: {ciclo}")
        total = sum(
            float(c["cantidad"]) * precio_de(c["ref"], (*visitando, codigo))
            for c in composicion
        )
        ficha["precio"] = round(total, 4)
        return ficha["precio"]

    for codigo in list(plano):
        precio_de(codigo)
    return plano


def _desglosar(ref: str, rendimiento: float, catalogo: dict, origen: str = "") -> list[dict]:
    """Devuelve las líneas físicas de un recurso, abriendo los compuestos.

    Un mortero elaborado en obra no se escribe como una línea opaca: se abre
    en el cemento, la arena y el agua que realmente lo componen. Es la forma
    en que se lee un análisis de precio unitario en Venezuela y, además, es lo
    que permite que al cambiar el precio del cemento en la aplicación se
    recalculen todas las partidas que llevan mortero.
    """
    ficha = catalogo.get(ref)
    if not ficha:
        raise ValueError(f"el recurso «{ref}» no existe en recursos.json")
    composicion = ficha.get("composicion")
    if not composicion:
        return [{
            "grupo": ficha["grupo"],
            "codigo": ref,
            "unidad": ficha["unidad"],
            "descripcion": ficha["descripcion"],
            "rendimiento": rendimiento,
            "precio": float(ficha["precio"]),
            "origen": origen,
        }]
    etiqueta = origen or _etiqueta_compuesto(ficha["descripcion"])
    lineas: list[dict] = []
    for componente in composicion:
        lineas.extend(_desglosar(
            componente["ref"],
            rendimiento * float(componente["cantidad"]),
            catalogo,
            etiqueta,
        ))
    return lineas


def _etiqueta_compuesto(descripcion: str) -> str:
    """«Mortero de pega para mampostería, elaborado…» -> «mortero de pega»."""
    corte = descripcion.split(",")[0].split(" para ")[0].strip().rstrip(".")
    return corte[:1].lower() + corte[1:] if corte else "mezcla en obra"


def resolver_recursos(partida: dict, catalogo: dict) -> list[dict]:
    """Convierte las líneas de la partida en recursos completos.

    Una línea puede venir de dos formas:
      * `{"ref": "MO-OF1-SOL", "rendimiento": 0.40}`  -> hereda grupo, unidad,
        descripción y precio del cuadro de recursos (recomendado).
      * `{"grupo": ..., "codigo": ..., "unidad": ..., "descripcion": ...,
         "rendimiento": ..., "precio": ...}` -> definida por completo en línea.
    """
    resueltos = []
    for linea in partida.get("recursos", []):
        ref = linea.get("ref")
        if ref:
            ficha = catalogo.get(ref)
            if not ficha:
                raise ValueError(
                    f"{partida['codigo']}: el recurso «{ref}» no existe en recursos.json"
                )
            if ficha.get("composicion") and not linea.get("sin_desglosar"):
                resueltos.extend(
                    _desglosar(ref, float(linea["rendimiento"]), catalogo)
                )
                continue
            resueltos.append({
                "grupo": ficha["grupo"],
                "codigo": ref,
                "unidad": linea.get("unidad") or ficha["unidad"],
                "descripcion": linea.get("descripcion") or ficha["descripcion"],
                "rendimiento": float(linea["rendimiento"]),
                "precio": float(linea.get("precio", ficha["precio"])),
            })
        else:
            resueltos.append({**linea, "rendimiento": float(linea["rendimiento"]),
                              "precio": float(linea["precio"])})

    # Un mismo componente puede llegar por dos caminos (p. ej. el cemento del
    # mortero de pega y el del friso en la misma partida). Se acumulan en una
    # sola línea para que el descompuesto no repita el mismo código.
    fusionadas: dict[tuple, dict] = {}
    orden: list[tuple] = []
    for item in resueltos:
        clave = (item["grupo"], item["codigo"], item["unidad"], round(item["precio"], 6))
        if clave in fusionadas:
            fusionadas[clave]["rendimiento"] += item["rendimiento"]
            previo = fusionadas[clave].get("origen") or ""
            nuevo = item.get("origen") or ""
            if nuevo and nuevo not in previo:
                fusionadas[clave]["origen"] = f"{previo} y {nuevo}" if previo else nuevo
        else:
            fusionadas[clave] = dict(item)
            orden.append(clave)

    salida = []
    for clave in orden:
        item = fusionadas[clave]
        item["rendimiento"] = round(item["rendimiento"], 8)
        origen = item.pop("origen", "")
        if origen:
            item["descripcion"] = f"{item['descripcion'].rstrip('.')}. Para {origen}."
        salida.append(item)
    return salida


def construir_hoja(partida: dict, ruta: Path) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = partida["codigo"][:31]

    negrita = Font(bold=True)
    ajuste = Alignment(wrap_text=True, vertical="top")
    fino = Side(style="thin", color="999999")

    # --- Etiquetas de clasificación (filas 1 y 2) ---
    # Van encima de la partida. Subcapítulo y apartado comparten la fila 2 en
    # celdas distintas para que ambos queden antes de la cabecera de la partida
    # y el importador pueda recuperarlos sin confundirlos con ella.
    ubicacion = partida.get("_ubicacion") or {}
    if ubicacion.get("capitulo"):
        ws.cell(1, 1, f"Capítulo: {ubicacion['capitulo_cod']} {ubicacion['capitulo']}")
    if ubicacion.get("subcapitulo"):
        ws.cell(
            2, 1,
            f"Subcapítulo: {ubicacion['capitulo_cod']}.{ubicacion['subcapitulo_cod']} "
            f"{ubicacion['subcapitulo']}",
        )
    if ubicacion.get("apartado"):
        ws.cell(
            2, 5,
            f"Apartado: {ubicacion['capitulo_cod']}.{ubicacion['subcapitulo_cod']}."
            f"{ubicacion['apartado_cod']} {ubicacion['apartado']}",
        )

    # --- Cabecera de la partida (filas 3 y 5) ---
    ws.cell(3, 1, partida["codigo"]).font = negrita
    ws.cell(3, 2, partida["unidad"])
    ws.cell(3, 3, partida["titulo"]).font = negrita
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=8)
    ws.cell(5, 1, partida["descripcion"].rstrip() + "\n").alignment = ajuste
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)

    # --- Fila de encabezados (fila 8) ---
    encabezados = {
        COL_CODIGO: "Código", COL_UNIDAD: "Unidad", COL_DESCRIPCION: "Descripción",
        COL_RENDIMIENTO: "Rendimiento", COL_PRECIO: "Precio\nunitario", COL_IMPORTE: "Importe",
    }
    for col, texto in encabezados.items():
        celda = ws.cell(8, col, texto)
        celda.font = negrita
        celda.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        celda.border = Border(top=fino, bottom=fino)

    fila = 9
    indice = 0
    filas_subtotal: list[int] = []   # filas donde queda cada subtotal
    totales: dict[str, float] = {}

    recursos = partida["_recursos"]
    for clave, etiqueta in ORDEN_GRUPOS:
        del_grupo = [r for r in recursos if r.get("grupo") == clave]
        if not del_grupo:
            continue
        indice += 1
        ws.cell(fila, COL_CODIGO, indice)
        ws.cell(fila, COL_DESCRIPCION, etiqueta).font = negrita
        fila += 1
        primera = fila
        acumulado = 0.0
        for r in del_grupo:
            rendimiento = float(r["rendimiento"])
            precio = float(r["precio"])
            acumulado += round(rendimiento * precio, 2)
            ws.cell(fila, COL_CODIGO, r["codigo"])
            ws.cell(fila, COL_UNIDAD, r["unidad"])
            ws.cell(fila, COL_DESCRIPCION, r["descripcion"]).alignment = ajuste
            ws.cell(fila, COL_RENDIMIENTO, rendimiento)
            ws.cell(fila, COL_PRECIO, precio)
            ws.cell(fila, COL_IMPORTE, round(rendimiento * precio, 2))
            fila += 1
        ws.cell(fila, COL_RENDIMIENTO, f"Subtotal {etiqueta.lower()}:").font = negrita
        ws.cell(fila, COL_IMPORTE, round(acumulado, 2))
        filas_subtotal.append(fila)
        totales[clave] = round(acumulado, 2)
        fila += 1

    if not filas_subtotal:
        raise ValueError(f"{partida['codigo']}: la partida no tiene recursos.")

    # --- Costes directos complementarios (% sobre la suma de subtotales) ---
    pct = float(partida.get("complementarios_pct", 2))
    indice += 1
    ws.cell(fila, COL_CODIGO, indice)
    ws.cell(fila, COL_DESCRIPCION, "Costes directos complementarios").font = negrita
    fila += 1
    fila_pct = fila
    ws.cell(fila, COL_UNIDAD, "%")
    ws.cell(fila, COL_DESCRIPCION, "Costes directos complementarios")
    ws.cell(fila, COL_RENDIMIENTO, pct)
    base = round(sum(totales.values()), 2)
    ws.cell(fila, COL_PRECIO, base)
    totales["complementarios"] = round(base * pct / 100, 2)
    ws.cell(fila, COL_IMPORTE, totales["complementarios"])
    fila += 1

    # --- Costes directos (1+2+...) ---
    fila_total = fila
    etiqueta_total = "Costes directos (" + "+".join(str(i) for i in range(1, indice + 1)) + "):"
    ws.cell(fila, COL_RENDIMIENTO, etiqueta_total).font = negrita
    ws.cell(fila, COL_IMPORTE, round(sum(totales.values()), 2)).font = negrita
    for col in range(1, 9):
        ws.cell(fila, col).border = Border(top=fino)

    # --- Formato ---
    for col, ancho in zip("ABCDEFGH", (16, 8, 4, 10, 62, 13, 12, 12)):
        ws.column_dimensions[col].width = ancho
    ws.row_dimensions[5].height = 60

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)

    total = round(sum(totales.values()), 2)
    return {"totales": totales, "coste_directo": total}


def validar(ruta: Path) -> dict:
    """Comprueba el .xlsx con el lector CYPE real del proyecto."""
    sys.path.insert(0, str(RAIZ))
    from app.services import importer

    datos = ruta.read_bytes()
    if not importer.es_formato_cype_xlsx(datos):
        return {"detectado": False}
    analisis = importer.analizar_cype_xlsx(datos)
    partidas = analisis.get("partidas", [])
    return {"detectado": True, "analisis": analisis, "partidas": partidas}


def main(argv: list[str]) -> int:
    if not ORIGEN.exists():
        sys.exit(f"No existe {ORIGEN}")
    fuentes = sorted(ORIGEN.glob("*.json"))
    if argv:
        fuentes = [f for f in fuentes if f.stem in argv]
    if not fuentes:
        sys.exit("No hay descompuestos que generar.")

    catalogo_recursos = cargar_recursos()
    taxonomia = cargar_clasificacion()
    n_caps = len(taxonomia.get("capitulos", {}))
    print(f"Cuadro de recursos: {len(catalogo_recursos)} recursos · "
          f"Clasificación: {n_caps} capítulos · Moneda: {taxonomia.get('_moneda', 'USD')}\n")

    fallos = 0
    resumen_catalogo: list[dict] = []
    for fuente in fuentes:
        partida = json.loads(fuente.read_text(encoding="utf-8"))
        partida["_recursos"] = resolver_recursos(partida, catalogo_recursos)
        partida["_ubicacion"] = ubicar(partida, taxonomia)
        destino = SALIDA / f"{partida['codigo']}.xlsx"
        resumen = construir_hoja(partida, destino)
        u = partida["_ubicacion"]
        print(f"\n{partida['codigo']}  {partida['titulo']}")
        print(
            f"  clasificación  : {u['capitulo']} › {u['subcapitulo']} › {u['apartado']}"
        )
        print(f"  archivo        : {destino.relative_to(RAIZ)}")
        for clave, valor in resumen["totales"].items():
            print(f"  {clave:<16}: {valor:>8.2f} USD")
        print(f"  {'COSTE DIRECTO':<16}: {resumen['coste_directo']:>8.2f} USD")

        v = validar(destino)
        if not v["detectado"]:
            print("  ✗ El proyecto NO lo reconoce como formato de descompuesto")
            fallos += 1
            continue
        leidas = v["partidas"]
        if not leidas:
            print("  ✗ Reconocido, pero sin partidas legibles")
            fallos += 1
            continue
        p = leidas[0]
        costes = p.get("costes", {})
        print(f"  ✓ detectado: código «{p.get('codigo')}», unidad «{p.get('unidad')}», "
              f"{len(p.get('filas', []))} filas")
        print(f"    costes leídos por la app: " + ", ".join(
            f"{k}={v:.2f}" for k, v in costes.items() if isinstance(v, (int, float))))
        horas = sum(r["rendimiento"] for r in partida["_recursos"]
                    if r["grupo"] == "mano_obra")
        print(f"    horas de mano de obra por {partida['unidad']}: {horas:.3f} h")

        pc = partida.get("producto_cliente")
        if pc:
            print(f"  producto cliente: {pc['consumo']} {pc['unidad']}/{partida['unidad']} "
                  f"— {pc['tipo'][:52]}")

        margen = float(partida.get("margen", MARGEN_DEFECTO))
        coste = resumen["coste_directo"]
        resumen_catalogo.append({
            "codigo": partida["codigo"],
            "codigo_legacy": partida.get("codigo_legacy", ""),
            "ubicacion": partida["_ubicacion"],
            "producto_cliente": pc,
            "capitulo": partida["_ubicacion"]["capitulo"],
            "titulo": partida["titulo"],
            "descripcion": partida["descripcion"],
            "unidad": partida["unidad"],
            "coste_directo": coste,
            "margen": margen,
            "precio_venta": round(coste * (1 + margen), 2),
            "horas": round(horas, 3),
            "costes": resumen["totales"],
        })

    if resumen_catalogo:
        escribir_catalogo(resumen_catalogo)
        escribir_arbol(resumen_catalogo, taxonomia)
    return 1 if fallos else 0


def escribir_catalogo(filas: list[dict]) -> None:
    """Vuelca el resumen de todas las partidas al maestro del catálogo."""
    ruta = BASE / "datos" / "partidas.csv"
    cabeceras = [
        "codigo", "codigo_legacy", "capitulo", "partida", "descripcion", "unidad", "precio",
        "categoria", "subcategoria", "apartado", "coste_materiales", "coste_mano_obra",
        "coste_complementarios", "coste_otros", "rendimiento",
        "desperdicio_pct", "notas_tecnicas",
    ]
    import csv as _csv
    with ruta.open("w", encoding="utf-8", newline="") as fh:
        w = _csv.writer(fh, delimiter=";", lineterminator="\n")
        w.writerow(cabeceras)
        for f in filas:
            c = f["costes"]
            u = f["ubicacion"]
            nota = f"Coste directo {f['coste_directo']:.2f} USD + margen {f['margen']*100:.0f}%"
            if f.get("producto_cliente"):
                pcl = f["producto_cliente"]
                nota += (f" | NO INCLUYE el producto de elección del cliente: "
                         f"{pcl['tipo']} ({pcl['consumo']} {pcl['unidad']}/{f['unidad']})")
            w.writerow([
                f["codigo"], f.get("codigo_legacy", ""), u["capitulo"],
                f["titulo"], f["descripcion"],
                {"m²": "m2", "m³": "m3"}.get(f["unidad"], f["unidad"]),
                f"{f['precio_venta']:.2f}",
                f"{u['capitulo_cod']} {u['capitulo']}",
                f"{u['capitulo_cod']}.{u['subcapitulo_cod']} {u['subcapitulo']}",
                f"{u['capitulo_cod']}.{u['subcapitulo_cod']}.{u['apartado_cod']} {u['apartado']}",
                f"{c.get('materiales', 0):.2f}",
                f"{c.get('mano_obra', 0):.2f}",
                f"{c.get('complementarios', 0):.2f}",
                f"{c.get('maquinaria', 0):.2f}",
                f"{f['horas']:.3f} h/{f['unidad']}",
                "", nota,
            ])
    print(f"\nCatálogo consolidado -> {ruta.relative_to(RAIZ)} ({len(filas)} partidas)")


def escribir_arbol(filas: list[dict], taxonomia: dict) -> None:
    """Genera capítulo → subcapítulo → apartado → partida."""
    caps = taxonomia.get("capitulos", {})
    arbol = []
    for cod_cap, cap in caps.items():
        nodo_subs = []
        for cod_sub, sub in cap["subcapitulos"].items():
            nodo_apartados = []
            for cod_apartado, nombre_apartado in sub.get("apartados", {}).items():
                hijas = [
                    {
                        "codigo": f["codigo"],
                        "codigo_legacy": f.get("codigo_legacy", ""),
                        "titulo": f["titulo"],
                        "unidad": f["unidad"],
                        "precio": f["precio_venta"],
                        "horas": f["horas"],
                        "producto_cliente": bool(f.get("producto_cliente")),
                    }
                    for f in filas
                    if f["ubicacion"]["capitulo_cod"] == cod_cap
                    and f["ubicacion"]["subcapitulo_cod"] == cod_sub
                    and f["ubicacion"]["apartado_cod"] == cod_apartado
                ]
                nodo_apartados.append({
                    "codigo": f"{cod_cap}.{cod_sub}.{cod_apartado}",
                    "nombre": nombre_apartado,
                    "partidas": hijas,
                    "n": len(hijas),
                })
            nodo_subs.append({
                "codigo": f"{cod_cap}.{cod_sub}",
                "nombre": sub["nombre"],
                "apartados": nodo_apartados,
                "n": sum(a["n"] for a in nodo_apartados),
            })
        arbol.append({
            "codigo": cod_cap,
            "nombre": cap["nombre"],
            "subcapitulos": nodo_subs,
            "n": sum(s["n"] for s in nodo_subs),
        })

    ruta = BASE / "salida" / "arbol_catalogo.json"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    ruta.write_text(json.dumps({
        "version": taxonomia.get("_version", 2),
        "formato_codigo": taxonomia.get("_formato_codigo", "CC.SS.AA.NNN"),
        "ambito": taxonomia.get("_ambito", "reforma"),
        "moneda": taxonomia.get("_moneda", "USD"),
        "arbol": arbol,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    con = sum(1 for c in arbol if c["n"])
    n_apartados = sum(
        len(s["apartados"]) for c in arbol for s in c["subcapitulos"]
    )
    print(
        f"Árbol de navegación   -> {ruta.relative_to(RAIZ)} "
        f"({len(arbol)} capítulos, {con} con partidas, {n_apartados} apartados)"
    )


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
