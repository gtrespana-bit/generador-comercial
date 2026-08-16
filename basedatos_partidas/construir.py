#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Constructor de la base de datos de partidas (trabajo EXTERNO al proyecto).

Lee el maestro `datos/partidas.csv` (fuente de verdad, editable a mano) y
genera en `salida/` los ficheros listos para subir al Generador Comercial:

  * catalogo_partidas.csv   -> asistente de importación (destino=catalogo)
  * catalogo_partidas.xlsx  -> mismo contenido en Excel
  * catalogo_partidas.json  -> volcado íntegro (costes desglosados, notas)

No importa ni modifica ningún módulo del proyecto salvo para VALIDAR
(lectura de app/services/importer.py), de modo que lo que se genera es
exactamente lo que la aplicación va a detectar.

Uso:
    python3 basedatos_partidas/construir.py
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent
RAIZ = BASE.parent
ORIGEN = BASE / "datos" / "partidas.csv"
SALIDA = BASE / "salida"

# Cabeceras que el importador reconoce automáticamente (ALIAS_CAMPOS).
CABECERAS_IMPORT = [
    "Capítulo",
    "Partida",
    "Descripción",
    "Unidad",
    "Cantidad",
    "Precio unitario",
    "Categoría",
    "Tipo",
]

# Unidades sin aviso en el validador (importer.UNIDADES_COMUNES).
UNIDADES_OK = {"ud", "m2", "m", "ml", "m3", "juego", "hora", "glb", "kg"}

# Normalización de unidades escritas "bonitas" a las que acepta el proyecto.
EQUIV_UNIDADES = {
    "m²": "m2", "M2": "m2", "M²": "m2",
    "m³": "m3", "M3": "m3", "M³": "m3",
    "u": "ud", "und": "ud", "uds": "ud", "UD": "ud", "Ud": "ud",
    "h": "hora", "hr": "hora", "Kg": "kg", "KG": "kg", "ML": "ml", "Ml": "ml",
    "pa": "glb", "PA": "glb", "partida alzada": "glb",
}


def num(valor: str, defecto: float = 0.0) -> float:
    texto = str(valor or "").strip().replace(" ", "").replace("€", "")
    if not texto:
        return defecto
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".") if texto.rfind(",") > texto.rfind(".") else texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return defecto


def leer_maestro() -> list[dict]:
    if not ORIGEN.exists():
        sys.exit(f"No existe el maestro: {ORIGEN}")
    with ORIGEN.open(encoding="utf-8-sig", newline="") as fh:
        filas = list(csv.DictReader(fh, delimiter=";"))
    limpias = []
    for i, fila in enumerate(filas, start=2):
        fila = {(k or "").strip(): (v or "").strip() for k, v in fila.items()}
        if not fila.get("partida"):
            continue
        unidad = fila.get("unidad", "ud")
        unidad = EQUIV_UNIDADES.get(unidad, unidad).lower() or "ud"
        fila["unidad"] = unidad
        fila["_linea"] = i
        limpias.append(fila)
    return limpias


def revisar(filas: list[dict]) -> list[str]:
    """Controles de calidad propios de la base de datos."""
    problemas: list[str] = []
    vistos: dict[str, int] = {}
    codigos: dict[str, int] = {}
    for fila in filas:
        linea, nombre = fila["_linea"], fila["partida"]
        clave = nombre.strip().lower()
        if clave in vistos:
            problemas.append(f"L{linea}: nombre duplicado «{nombre}» (ya en L{vistos[clave]}). El catálogo lo omitiría.")
        vistos[clave] = linea

        codigo = fila.get("codigo", "").strip().upper()
        if codigo:
            if codigo in codigos:
                problemas.append(f"L{linea}: código duplicado «{codigo}» (ya en L{codigos[codigo]}).")
            codigos[codigo] = linea

        if len(nombre) > 200:
            problemas.append(f"L{linea}: el nombre supera 200 caracteres (columna nombre del modelo Partida).")
        if fila["unidad"] not in UNIDADES_OK:
            problemas.append(f"L{linea}: unidad «{fila['unidad']}» generará aviso en el importador.")

        precio = num(fila.get("precio"))
        if precio <= 0:
            problemas.append(f"L{linea}: precio 0 o vacío en «{nombre}».")
        desglose = sum(num(fila.get(c)) for c in
                       ("coste_materiales", "coste_mano_obra", "coste_complementarios", "coste_otros"))
        if desglose and precio and desglose > precio * 1.001:
            problemas.append(f"L{linea}: el desglose de costes ({desglose:.2f}) supera al precio ({precio:.2f}).")
        if not fila.get("descripcion"):
            problemas.append(f"L{linea}: «{nombre}» sin descripción.")
        if not fila.get("capitulo"):
            problemas.append(f"L{linea}: «{nombre}» sin capítulo.")
    return problemas


def validar_con_el_proyecto(ruta_csv: Path) -> dict:
    """Pasa el CSV generado por el importador REAL del proyecto."""
    sys.path.insert(0, str(RAIZ))
    from app.services import importer  # noqa: E402

    datos = ruta_csv.read_bytes()
    matriz = importer.leer_csv(datos)
    analisis = importer.analizar_matriz(matriz, tiene_encabezados=True)
    mapeo = importer.detectar_mapeo(analisis["encabezados"])
    resultado = importer.validar_filas(analisis["filas"], mapeo)
    return {
        "campos_detectados": {k: v for k, v in mapeo.items() if v is not None},
        "campos_sin_detectar": [k for k, v in mapeo.items() if v is None],
        "filas_ok": len(resultado["filas"]),
        "errores": resultado["errores"],
        "advertencias": resultado["advertencias"],
    }


def escribir_csv(filas: list[dict]) -> Path:
    ruta = SALIDA / "catalogo_partidas.csv"
    with ruta.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(CABECERAS_IMPORT)
        for f in filas:
            w.writerow([
                f.get("capitulo", ""),
                f.get("partida", ""),
                f.get("descripcion", ""),
                f.get("unidad", "ud"),
                "1",
                f"{num(f.get('precio')):.2f}".replace(".", ","),
                f.get("categoria") or f.get("capitulo", "General"),
                "incluida",
            ])
    return ruta


def escribir_xlsx(filas: list[dict]) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  ! openpyxl no instalado: se omite el .xlsx (pip install openpyxl)")
        return None
    ruta = SALIDA / "catalogo_partidas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    ws.append(CABECERAS_IMPORT)
    cabecera = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="1F3864")
    for celda in ws[1]:
        celda.font, celda.fill = cabecera, relleno
    for f in filas:
        ws.append([
            f.get("capitulo", ""),
            f.get("partida", ""),
            f.get("descripcion", ""),
            f.get("unidad", "ud"),
            1,
            round(num(f.get("precio")), 2),
            f.get("categoria") or f.get("capitulo", "General"),
            "incluida",
        ])
    for col, ancho in zip("ABCDEFGH", (22, 55, 90, 10, 10, 14, 20, 12)):
        ws.column_dimensions[col].width = ancho
    for fila in ws.iter_rows(min_row=2):
        fila[2].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(ruta)
    return ruta


def escribir_json(filas: list[dict]) -> Path:
    ruta = SALIDA / "catalogo_partidas.json"
    payload = []
    for f in filas:
        payload.append({
            "codigo": f.get("codigo", ""),
            "capitulo": f.get("capitulo", ""),
            "nombre": f.get("partida", ""),
            "descripcion": f.get("descripcion", ""),
            "unidad": f.get("unidad", "ud"),
            "precio_unitario": round(num(f.get("precio")), 2),
            "categoria": f.get("categoria") or f.get("capitulo", "General"),
            "subcategoria": f.get("subcategoria", ""),
            "coste_materiales": round(num(f.get("coste_materiales")), 2),
            "coste_mano_obra": round(num(f.get("coste_mano_obra")), 2),
            "coste_complementarios": round(num(f.get("coste_complementarios")), 2),
            "coste_otros": round(num(f.get("coste_otros")), 2),
            "rendimiento": f.get("rendimiento", ""),
            "desperdicio_recomendado_pct": num(f.get("desperdicio_pct")),
            "notas_tecnicas": f.get("notas_tecnicas", ""),
        })
    ruta.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return ruta


def main() -> int:
    SALIDA.mkdir(parents=True, exist_ok=True)
    filas = leer_maestro()
    print(f"Maestro: {len(filas)} partidas leídas de {ORIGEN.relative_to(RAIZ)}")

    problemas = revisar(filas)
    if problemas:
        print(f"\nRevisión de calidad ({len(problemas)} avisos):")
        for p in problemas[:40]:
            print("  -", p)
        if len(problemas) > 40:
            print(f"  ... y {len(problemas) - 40} más")
    else:
        print("Revisión de calidad: sin incidencias.")

    csv_out = escribir_csv(filas)
    xlsx_out = escribir_xlsx(filas)
    json_out = escribir_json(filas)
    print("\nGenerado:")
    for ruta in (csv_out, xlsx_out, json_out):
        if ruta:
            print("  ", ruta.relative_to(RAIZ))

    print("\nValidación con el importador real del proyecto:")
    v = validar_con_el_proyecto(csv_out)
    print("   campos detectados :", ", ".join(v["campos_detectados"]))
    if v["campos_sin_detectar"]:
        print("   sin detectar      :", ", ".join(v["campos_sin_detectar"]))
    print("   filas válidas     :", v["filas_ok"])
    print("   errores           :", len(v["errores"]))
    print("   advertencias      :", len(v["advertencias"]))
    for aviso in v["errores"][:10]:
        print("     ERROR fila", aviso.get("fila"), "-", aviso.get("mensaje"))
    for aviso in v["advertencias"][:10]:
        print("     aviso fila", aviso.get("fila"), "-", aviso.get("mensaje"))
    return 1 if v["errores"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
