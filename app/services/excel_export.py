"""Servicio de exportación a Excel con formato profesional.

Genera archivos .xlsx con:
- Formato condicional (totales en negrita, filas alternas)
- Múltiples hojas (resumen, detalle, costes)
- Columnas con ancho optimizado
- Títulos y subtítulos
"""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Estilos compartidos
# ---------------------------------------------------------------------------

_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="0D9488")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="0F172A")
_SUBTITLE_FONT = Font(bold=True, size=11, color="64748B")
_TOTAL_FILL = PatternFill("solid", fgColor="F0FDFA")
_TOTAL_FONT = Font(bold=True, size=11, color="0D9488")
_ODD_FILL = PatternFill("solid", fgColor="F8FAFC")
_CURRENCY_FMT = '#,##0.00'  # La moneda se escribe en la columna Moneda; nunca fijar USD
_PCT_FMT = '0.00"%"'


def _style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _style_total_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER


def _auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = max(length, len(str(cell.value)))
        adjusted = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _money(valor):
    try:
        return round(float(valor or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _fmt_monto(valor):
    """Formatear monto como texto con formato venezolano."""
    v = _money(valor)
    return f"${v:,.2f}"


# ---------------------------------------------------------------------------
# Exportar presupuesto a Excel
# ---------------------------------------------------------------------------

def exportar_presupuesto_excel(presupuesto, cfg=None):
    """Genera un libro Excel con múltiples hojas para un presupuesto.

    Hojas:
      1. Resumen — datos generales del presupuesto
      2. Detalle — capítulos y partidas con totales
      3. Costes — desglose de costes internos (si existen)
    """
    wb = Workbook()

    # --- Hoja 1: Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.sheet_view.showGridLines = False

    _write_resumen_sheet(ws_resumen, presupuesto, cfg)

    # --- Hoja 2: Detalle ---
    ws_detalle = wb.create_sheet("Detalle")
    ws_detalle.sheet_view.showGridLines = False
    _write_detalle_sheet(ws_detalle, presupuesto)

    # --- Hoja 3: Costes (si corresponde) ---
    if hasattr(presupuesto, "todas_partidas"):
        tiene_costes = any(
            (p.coste_materiales or 0) + (p.coste_mano_obra or 0) +
            (p.coste_complementarios or 0) + (p.coste_otros or 0) > 0
            for p in presupuesto.todas_partidas
        )
        if tiene_costes:
            ws_costes = wb.create_sheet("Costes Internos")
            ws_costes.sheet_view.showGridLines = False
            _write_costes_sheet(ws_costes, presupuesto)

    # Ajustar anchos
    for ws in [ws_resumen, ws_detalle]:
        _auto_width(ws)
    if "Costes Internos" in wb.sheetnames:
        _auto_width(wb["Costes Internos"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_resumen_sheet(ws, presupuesto, cfg):
    """Hoja de resumen del presupuesto."""
    ws["A1"] = "RESUMEN DEL PRESUPUESTO"
    ws["A1"].font = _TITLE_FONT

    ws["A3"] = "Datos del documento"
    ws["A3"].font = _SUBTITLE_FONT

    datos = [
        ("Número", presupuesto.numero),
        ("Fecha", presupuesto.fecha.isoformat() if presupuesto.fecha else ""),
        ("Estado", presupuesto.estado or ""),
        ("Título / Obra", presupuesto.titulo or ""),
        ("Dirección", presupuesto.direccion_obra or ""),
        ("Código postal", presupuesto.codigo_postal or ""),
        ("Cliente", presupuesto.cliente.nombre if presupuesto.cliente else ""),
        ("Moneda", presupuesto.moneda or "USD"),
    ]

    if presupuesto.tipo_cambio:
        datos.append(("Tasa de cambio (USD → moneda contractual)", f"{presupuesto.tipo_cambio:.2f} {presupuesto.moneda or 'USD'}"))
    if presupuesto.fecha_tipo_cambio:
        datos.append(("Fecha tasa", presupuesto.fecha_tipo_cambio.isoformat()))

    datos.extend([
        ("IVA (%)", presupuesto.impuesto_pct),
        ("Descuento (%)", presupuesto.descuento_pct),
        ("Validez (días)", presupuesto.validez_dias),
    ])

    row = 4
    for label, valor in datos:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F1F5F9")
        ws.cell(row=row, column=1).border = _BORDER
        cell_valor = ws.cell(row=row, column=2, value=valor)
        cell_valor.border = _BORDER
        if isinstance(valor, float):
            cell_valor.number_format = '#,##0.00'
        row += 1

    # Totales
    row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = _SUBTITLE_FONT
    row += 1

    totales = [
        ("Subtotal (incluido)", _money(getattr(presupuesto, "subtotal", 0))),
        ("Subtotal opcionales", _money(getattr(presupuesto, "subtotal_opcional", 0))),
        ("Subtotal alternativas", _money(getattr(presupuesto, "subtotal_alternativas", 0))),
    ]
    if getattr(presupuesto, "total_productos", 0):
        totales.extend([
            ("Total obra / partidas", _money(getattr(presupuesto, "subtotal_obra", 0))),
            ("Total productos", _money(getattr(presupuesto, "total_productos", 0))),
        ])
    totales.extend([
        ("Costes adicionales", _money(getattr(presupuesto, "costes_adicionales", 0))),
        ("Descuento", -_money(getattr(presupuesto, "descuento_monto", 0))),
        ("Base imponible", _money(getattr(presupuesto, "base", 0))),
        ("IVA", _money(getattr(presupuesto, "impuesto_monto", 0))),
        ("TOTAL", _money(getattr(presupuesto, "total", 0))),
    ])

    for label, valor in totales:
        ws.cell(row=row, column=1, value=label).border = _BORDER
        c = ws.cell(row=row, column=2, value=valor)
        c.number_format = _CURRENCY_FMT
        c.border = _BORDER
        if label == "TOTAL":
            _style_total_row(ws, row, 2)
        row += 1

    # Márgenes si existen
    if hasattr(presupuesto, "coste_interno") and presupuesto.coste_interno:
        row += 1
        filas_margen = [
            ("Coste interno", _money(presupuesto.coste_interno), _CURRENCY_FMT),
        ]
        if getattr(presupuesto, "total_productos", 0):
            filas_margen.extend([
                ("Beneficio obra", _money(getattr(presupuesto, "margen_obra", 0)), _CURRENCY_FMT),
                ("Margen obra (%)", _money(getattr(presupuesto, "margen_obra_pct", 0)), '0.00"%"'),
                ("Beneficio productos", _money(getattr(presupuesto, "margen_productos", 0)), _CURRENCY_FMT),
                ("Margen productos (%)", _money(getattr(presupuesto, "margen_productos_pct", 0)), '0.00"%"'),
            ])
        filas_margen.extend([
            ("Beneficio total", _money(getattr(presupuesto, "margen", 0)), _CURRENCY_FMT),
            ("Margen total (%)", _money(getattr(presupuesto, "margen_pct", 0)), '0.00"%"'),
        ])
        for label, valor, fmt in filas_margen:
            ws.cell(row=row, column=1, value=label).border = _BORDER
            c = ws.cell(row=row, column=2, value=valor)
            c.number_format = fmt
            c.border = _BORDER
            row += 1

    # Notas y condiciones
    if presupuesto.notas:
        row += 2
        ws.cell(row=row, column=1, value="Información adicional").font = _SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value=presupuesto.notas)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 80

    if presupuesto.condiciones:
        row += 2
        ws.cell(row=row, column=1, value="Condiciones comerciales").font = _SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value=presupuesto.condiciones)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 80


def _money_round(valor):
    """Redondeo monetario ROUND_HALF_UP a 2 decimales, idéntico al motor de
    cálculos de la aplicación (calculations.py)."""
    from decimal import Decimal, ROUND_HALF_UP
    try:
        return float(Decimal(str(valor or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0.0


def _coste_partida_excel(p):
    """Coste interno de una partida con las MISMAS reglas que el servidor:
    descompuesto CYPE → coste directo + producto (sin desperdicio); resto →
    (materiales + mano de obra + complementarios + otros) × (1 + desperdicio)
    + producto. Redondeado en cada paso."""
    cantidad = _money_round(getattr(p, "cantidad_total", 0) or 0)
    descompuesto = getattr(p, "descomposicion_cype", None)
    producto_coste = _money_round(getattr(p, "producto_coste", 0) or 0)
    if descompuesto is not None and getattr(descompuesto, "coste_directo_unitario", None) is not None:
        directo = _money_round(descompuesto.coste_directo_unitario or 0)
        if getattr(descompuesto, "origen", "") != "manual":
            return _money_round(cantidad * (directo + producto_coste))
        desperdicio = _money_round(getattr(p, "desperdicio_pct", 0) or 0)
        return _money_round(cantidad * (directo * (1 + desperdicio / 100) + producto_coste))
    materiales = _money_round(getattr(p, "coste_materiales", 0) or 0)
    mano_obra = _money_round(getattr(p, "coste_mano_obra", 0) or 0)
    complementarios = _money_round(getattr(p, "coste_complementarios", 0) or 0)
    otros = _money_round(getattr(p, "coste_otros", 0) or 0)
    desperdicio = _money_round(getattr(p, "desperdicio_pct", 0) or 0)
    subtotal = materiales + mano_obra + complementarios + otros
    return _money_round(cantidad * (subtotal * (1 + desperdicio / 100) + producto_coste))


def _write_detalle_sheet(ws, presupuesto):
    """Hoja de detalle con capítulos y partidas."""
    ws["A1"] = f"Detalle del Presupuesto {getattr(presupuesto, 'numero', '')}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Capítulo", "Partida", "Descripción", "Unidad", "Cantidad",
               "Precio unit.", "Importe", "Coste interno", "Beneficio"]

    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    row = 4
    alt = False
    cap_fills = {}

    total_capitulo = 0
    total_gral = 0
    total_coste = 0

    for cap in getattr(presupuesto, "capitulos", []):
        cap_nombre = cap.nombre or "Sin capítulo"
        cap_subtotal = 0
        cap_coste = 0

        avanzadas = bool(getattr(presupuesto, "usar_funciones_avanzadas", False))
        for p in getattr(cap, "partidas", []):
            cantidad = getattr(p, "cantidad_total", 0) or 0
            precio = getattr(p, "precio_unitario", 0) or 0
            # Importe redondeado con la misma regla que filas, capítulos y
            # totales del presupuesto (p.importe usa ROUND_HALF_UP).
            importe = _money_round(cantidad * precio)

            coste_mat = getattr(p, "coste_materiales", 0) or 0
            coste_mo = getattr(p, "coste_mano_obra", 0) or 0
            coste_comp = getattr(p, "coste_complementarios", 0) or 0
            coste_otros = getattr(p, "coste_otros", 0) or 0
            desperdicio = getattr(p, "desperdicio_pct", 0) or 0

            coste_base = coste_mat + coste_mo + coste_comp + coste_otros
            # Mismo coste que el motor del servidor (incluye el coste del
            # producto asociado y respeta el coste directo CYPE).
            coste = _coste_partida_excel(p)
            beneficio = _money_round(importe - coste)

            # Solo las partidas que forman parte del total del documento
            # (mismo filtro que calculations.calcular_totales) entran en los
            # subtotales; las opcionales/alternativas no seleccionadas y las
            # excluidas se muestran pero no se suman.
            tipo = (getattr(p, "tipo_partida", "included") or "included").lower()
            if avanzadas and tipo in ("optional", "alternative") and not getattr(p, "seleccionada", False):
                incluir_en_total = False
            elif tipo == "excluded":
                incluir_en_total = False
            else:
                incluir_en_total = True

            ws.cell(row=row, column=1, value=cap_nombre)
            ws.cell(row=row, column=2, value=p.nombre or "")
            ws.cell(row=row, column=3, value=p.descripcion or "")
            ws.cell(row=row, column=4, value=p.unidad or "ud")
            ws.cell(row=row, column=5, value=cantidad)
            ws.cell(row=row, column=6, value=precio)
            ws.cell(row=row, column=7, value=importe)
            ws.cell(row=row, column=8, value=coste if coste_base > 0 else None)
            ws.cell(row=row, column=9, value=beneficio if coste_base > 0 else None)

            # Formato de celdas
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = _BORDER
                if col in (5, 6, 7, 8, 9):
                    cell.number_format = _CURRENCY_FMT
                    cell.alignment = Alignment(horizontal="right")

            # Fila alterna
            if alt:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = _ODD_FILL
            alt = not alt

            if incluir_en_total:
                cap_subtotal = _money_round(cap_subtotal + importe)
                cap_coste = _money_round(cap_coste + coste)
                total_gral = _money_round(total_gral + importe)
                total_coste = _money_round(total_coste + coste)
            row += 1

        # Subtotal capítulo (redondeado, coincide con el PDF)
        ws.cell(row=row, column=1, value=f"Subtotal {cap_nombre}")
        ws.cell(row=row, column=7, value=_money_round(cap_subtotal))
        ws.cell(row=row, column=7).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F1F5F9")
            ws.cell(row=row, column=col).border = _BORDER
        row += 1

    # Total general (redondeado, coincide con el resumen y el PDF)
    ws.cell(row=row, column=1, value="TOTAL INCLUIDO")
    ws.cell(row=row, column=7, value=_money_round(total_gral))
    ws.cell(row=row, column=7).number_format = _CURRENCY_FMT
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=8, value=total_coste)
    ws.cell(row=row, column=8).number_format = _CURRENCY_FMT
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=9, value=total_gral - total_coste)
    ws.cell(row=row, column=9).number_format = _CURRENCY_FMT
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
    _style_total_row(ws, row, len(headers))


def _write_costes_sheet(ws, presupuesto):
    """Hoja de desglose de costes internos por partida."""
    ws["A1"] = f"Costes Internos — Presupuesto {getattr(presupuesto, 'numero', '')}"
    ws["A1"].font = _TITLE_FONT

    headers = ["Capítulo", "Partida", "Coste materiales", "Coste mano obra",
               "Complementarios", "Otros costes", "Desperdicio %", "Coste total"]

    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    row = 4
    total_materiales = 0
    total_mano_obra = 0
    total_complementarios = 0
    total_otros = 0
    total_coste = 0

    avanzadas = bool(getattr(presupuesto, "usar_funciones_avanzadas", False))
    for cap in getattr(presupuesto, "capitulos", []):
        for p in getattr(cap, "partidas", []):
            coste_mat = getattr(p, "coste_materiales", 0) or 0
            coste_mo = getattr(p, "coste_mano_obra", 0) or 0
            coste_comp = getattr(p, "coste_complementarios", 0) or 0
            coste_otros = getattr(p, "coste_otros", 0) or 0
            desperdicio = getattr(p, "desperdicio_pct", 0) or 0

            coste_base = coste_mat + coste_mo + coste_comp + coste_otros
            # Mismo coste que el motor (incluye producto y respeta CYPE).
            coste = _coste_partida_excel(p)

            # Mismo filtro de inclusión que el motor de totales.
            tipo = (getattr(p, "tipo_partida", "included") or "included").lower()
            if avanzadas and tipo in ("optional", "alternative") and not getattr(p, "seleccionada", False):
                incluir_en_total = False
            elif tipo == "excluded":
                incluir_en_total = False
            else:
                incluir_en_total = True

            if coste_base > 0:
                ws.cell(row=row, column=1, value=cap.nombre or "")
                ws.cell(row=row, column=2, value=p.nombre or "")
                ws.cell(row=row, column=3, value=coste_mat)
                ws.cell(row=row, column=4, value=coste_mo)
                ws.cell(row=row, column=5, value=coste_comp)
                ws.cell(row=row, column=6, value=coste_otros)
                ws.cell(row=row, column=7, value=desperdicio)
                ws.cell(row=row, column=8, value=coste)

                for col in range(3, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = _CURRENCY_FMT if col != 7 else '0.00"%"'
                    cell.alignment = Alignment(horizontal="right")

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = _BORDER

                if incluir_en_total:
                    total_materiales = _money_round(total_materiales + coste_mat)
                    total_mano_obra = _money_round(total_mano_obra + coste_mo)
                    total_complementarios = _money_round(total_complementarios + coste_comp)
                    total_otros = _money_round(total_otros + coste_otros)
                    total_coste = _money_round(total_coste + coste)
                row += 1

    # Totales
    ws.cell(row=row, column=1, value="TOTALES")
    ws.cell(row=row, column=3, value=total_materiales)
    ws.cell(row=row, column=4, value=total_mano_obra)
    ws.cell(row=row, column=5, value=total_complementarios)
    ws.cell(row=row, column=6, value=total_otros)
    ws.cell(row=row, column=8, value=total_coste)
    for col in range(3, 9):
        ws.cell(row=row, column=col).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
    _style_total_row(ws, row, len(headers))


# ---------------------------------------------------------------------------
# Exportar lista de presupuestos (historial) a Excel
# ---------------------------------------------------------------------------

def exportar_historial_excel(presupuestos, cfg=None):
    """Exporta historic de presupuestos a Excel con formato profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Historial de Presupuestos"
    ws["A1"].font = _TITLE_FONT

    if cfg:
        ws["A2"] = f"Empresa: {cfg.empresa_nombre or ''}"
        ws["A2"].font = Font(italic=True, color="64748B")

    headers = ["Nº", "Fecha", "Cliente", "Título", "Estado", "Moneda",
               "Base imponible", "IVA", "Descuento", "TOTAL"]

    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    row = 5
    alt = False
    for p in presupuestos:
        ws.cell(row=row, column=1, value=p.numero)
        ws.cell(row=row, column=2, value=p.fecha.isoformat() if p.fecha else "")
        ws.cell(row=row, column=3, value=p.cliente.nombre if p.cliente else "")
        ws.cell(row=row, column=4, value=p.titulo or "")
        ws.cell(row=row, column=5, value=p.estado or "")
        ws.cell(row=row, column=6, value=p.moneda or "USD")

        base = _money(getattr(p, "base", 0))
        iva = _money(getattr(p, "impuesto_monto", 0))
        dto = _money(getattr(p, "descuento_monto", 0))
        total = _money(getattr(p, "total", 0))

        ws.cell(row=row, column=7, value=base)
        ws.cell(row=row, column=7).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=8, value=iva)
        ws.cell(row=row, column=8).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=9, value=dto)
        ws.cell(row=row, column=9).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=10, value=total)
        ws.cell(row=row, column=10).number_format = _CURRENCY_FMT

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.border = _BORDER
            if col >= 7:
                c.alignment = Alignment(horizontal="right")
            if alt:
                c.fill = _ODD_FILL
        alt = not alt
        row += 1

    # Totales
    ws.cell(row=row, column=1, value="TOTALES")
    ws.cell(row=row, column=7, value=sum(_money(getattr(p, "base", 0)) for p in presupuestos))
    ws.cell(row=row, column=8, value=sum(_money(getattr(p, "impuesto_monto", 0)) for p in presupuestos))
    ws.cell(row=row, column=9, value=sum(_money(getattr(p, "descuento_monto", 0)) for p in presupuestos))
    ws.cell(row=row, column=10, value=sum(_money(getattr(p, "total", 0)) for p in presupuestos))
    for col in range(7, 11):
        ws.cell(row=row, column=col).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
    _style_total_row(ws, row, len(headers))

    _auto_width(ws)
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["C"].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Exportar catálogo de partidas a Excel
# ---------------------------------------------------------------------------

def exportar_catalogo_partidas_excel(partidas):
    """Exporta el catálogo de partidas a Excel profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Partidas"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Catálogo de Partidas"
    ws["A1"].font = _TITLE_FONT

    ws["A2"] = f"Total: {len(partidas)} partidas"
    ws["A2"].font = _SUBTITLE_FONT

    headers = ["Código", "Código anterior", "Nombre", "Descripción", "Unidad",
               "Precio unit.", "Capítulo", "Subcapítulo", "Apartado",
               "Coste materiales", "Coste mano obra", "Complementarios", "Otros",
               "Tiempo est.", "Proveedor", "Rendimiento", "Desperdicio %",
               "Notas técnicas", "Última actualización", "Usos"]

    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    row = 5
    alt = False
    for p in partidas:
        valores = [
            p.codigo_interno or "",
            p.codigo_legacy or "",
            p.nombre or "",
            p.descripcion or "",
            p.unidad or "ud",
            _money(p.precio_unitario),
            p.categoria or "",
            p.subcategoria or "",
            p.apartado or "",
            _money(p.coste_materiales),
            _money(p.coste_mano_obra),
            _money(p.coste_complementarios),
            _money(p.coste_otros),
            p.tiempo_estimado_horas or "",
            p.proveedor or "",
            p.rendimiento or "",
            p.desperdicio_recomendado_pct or 0,
            p.notas_tecnicas or "",
            p.fecha_actualizacion_precio.isoformat() if p.fecha_actualizacion_precio else "",
            p.usos or 0,
        ]
        for col, valor in enumerate(valores, 1):
            ws.cell(row=row, column=col, value=valor)
        for col in (6, 10, 11, 12, 13):
            ws.cell(row=row, column=col).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=17).number_format = '0.00"%"'

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.border = _BORDER
            if col in (6, 10, 11, 12, 13, 17):
                c.alignment = Alignment(horizontal="right")
            if alt:
                c.fill = _ODD_FILL
        alt = not alt
        row += 1

    _auto_width(ws)
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["R"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Exportar catálogo de productos a Excel
# ---------------------------------------------------------------------------

def exportar_catalogo_productos_excel(productos):
    """Exporta el catálogo de productos a Excel profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Catálogo de Productos"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Total: {len(productos)} productos"
    ws["A2"].font = _SUBTITLE_FONT

    headers = ["Nombre", "Marca", "Modelo", "SKU", "Descripción", "Unidad",
               "Precio compra", "Precio venta", "Categoría", "Proveedor",
               "Color", "Acabado", "Formato", "Entrega (días)", "Variantes",
               "Ficha técnica", "Última actualización", "Usos"]

    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    row = 5
    alt = False
    for prod in productos:
        ws.cell(row=row, column=1, value=prod.nombre or "")
        ws.cell(row=row, column=2, value=prod.marca or "")
        ws.cell(row=row, column=3, value=prod.modelo or "")
        ws.cell(row=row, column=4, value=prod.sku or "")
        ws.cell(row=row, column=5, value=prod.descripcion or "")
        ws.cell(row=row, column=6, value=prod.unidad or "ud")
        if prod.precio_compra is not None:
            ws.cell(row=row, column=7, value=_money(prod.precio_compra))
            ws.cell(row=row, column=7).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=8, value=_money(prod.precio_unitario))
        ws.cell(row=row, column=8).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=9, value=prod.categoria or "")
        ws.cell(row=row, column=10, value=prod.proveedor or "")
        ws.cell(row=row, column=11, value=prod.color or "")
        ws.cell(row=row, column=12, value=prod.acabado or "")
        ws.cell(row=row, column=13, value=prod.formato or "")
        ws.cell(row=row, column=14, value=prod.tiempo_entrega_dias or "")
        ws.cell(row=row, column=15, value=prod.variantes or "")
        ws.cell(row=row, column=16, value=prod.ficha_tecnica or "")
        ws.cell(row=row, column=17, value=prod.fecha_actualizacion_precio.isoformat() if prod.fecha_actualizacion_precio else "")
        ws.cell(row=row, column=18, value=prod.usos or 0)

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.border = _BORDER
            if col in (7, 8, 14, 18):
                c.alignment = Alignment(horizontal="right")
            if alt:
                c.fill = _ODD_FILL
        alt = not alt
        row += 1

    _auto_width(ws)
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["O"].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Exportar catálogo de recursos (precios unitarios) a Excel
# ---------------------------------------------------------------------------

def exportar_catalogo_recursos_excel(recursos, etiquetas=None, moneda="USD", factor=1.0,
                                     precios_efectivos=None):
    """Exporta el catálogo central de recursos a Excel profesional.

    ``etiquetas`` mapea la categoría interna (mano_obra, materiales…) a su
    etiqueta legible; si no se pasa, se usa el valor interno.
    ``moneda``/``factor`` expresan el precio en la moneda de la organización
    (el catálogo interno se guarda en USD), con una columna que la explicita.
    ``precios_efectivos`` (opcional) es un diccionario ``id -> precio`` en
    ``moneda`` ya resuelto con la referencia nacional; si se pasa, tiene
    prioridad sobre ``precio base × factor`` para que exportar no devuelva el
    precio de la partida original convertido.
    """
    from ..services.tasa import tasa_convertir_precio

    etiquetas = etiquetas or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Recursos"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Catálogo de Recursos (Precios Unitarios)"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Total: {len(recursos)} recursos"
    ws["A2"].font = _SUBTITLE_FONT

    headers = ["Código", "Descripción", "Unidad", "Categoría", "Grupo",
               f"Precio ({moneda})", "Usos", "Proveedor", "Última actualización"]

    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    row = 5
    alt = False
    for r in recursos:
        ws.cell(row=row, column=1, value=r.codigo or "")
        ws.cell(row=row, column=2, value=r.descripcion or "")
        ws.cell(row=row, column=3, value=r.unidad or "")
        ws.cell(row=row, column=4, value=etiquetas.get(r.categoria, r.categoria) or "")
        ws.cell(row=row, column=5, value=r.grupo or "")
        _precio = (
            float(precios_efectivos.get(r.id))
            if precios_efectivos and precios_efectivos.get(r.id) is not None
            else tasa_convertir_precio(r.precio or 0, factor)
        )
        ws.cell(row=row, column=6, value=_money(_precio))
        ws.cell(row=row, column=6).number_format = _CURRENCY_FMT
        ws.cell(row=row, column=7, value=r.usos or 0)
        ws.cell(row=row, column=8, value=r.proveedor or "")
        ws.cell(row=row, column=9, value=r.fecha_actualizacion_precio.isoformat() if r.fecha_actualizacion_precio else "")

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.border = _BORDER
            if col in (6, 7):
                c.alignment = Alignment(horizontal="right")
            if alt:
                c.fill = _ODD_FILL
        alt = not alt
        row += 1

    _auto_width(ws)
    ws.column_dimensions["B"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
