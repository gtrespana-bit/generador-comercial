"""Lectura y validación de importaciones tabulares para presupuestos.

El navegador se ocupa de la experiencia de mapeo y vista previa; este módulo
mantiene en un solo lugar las reglas que no se deben confiar al cliente:
lectura de CSV/XLSX, detección de columnas, números con formato local y
validación de las filas antes de crear o ampliar un presupuesto.

Además del asistente tabular, reconoce el formato de descompuestos de CYPE
como ``DPT020.xlsx`` o ``RBE010c8_0_1_1c7_0_1_1c10_0_0.xlsx``. Ese formato no
es una tabla plana: contiene cabecera, filas de grupos, recursos, subtotales
y fórmulas. Por ello se lee en un flujo específico que conserva la posición
de *todas* las filas y columnas, las fórmulas y los rangos combinados; el
archivo .xlsx original se guarda en el servidor al confirmarlo.

Las columnas no ocupan posiciones fijas: el exportador de CYPE usa layouts
distintos según el descompuesto (p. ej. 8 columnas con «Unidad» en D, o
10 columnas con separadores estrechos y «Unidad» en C). La fila de
encabezados («Código, Unidad, Descripción, Rendimiento, Precio unitario,
Importe») se detecta dinámicamente y de ella se derivan las posiciones de
cada campo; así ambos formatos — y cualquier variante que respete esos
encabezados — se importan sin perder filas ni columnas. No todas las
partidas tienen gastos de materiales (las hay solo de mano de obra): los
grupos se clasifican igualmente en materiales, mano de obra, directos
complementarios y otros.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

MAX_FILAS = 1000
MAX_COLUMNAS = 30
MAX_BYTES = 8 * 1024 * 1024
# El formato CYPE contiene celdas vacías intencionadas y, en ocasiones, una
# zona de impresión de hasta 200 filas. Se conserva completa; nunca se corta
# silenciosamente. Si excede el límite se rechaza el archivo completo.
MAX_FILAS_CYPE = 5000
MAX_COLUMNAS_CYPE = 50
UNIDADES_COMUNES = {"ud", "m2", "m", "ml", "m3", "juego", "hora", "glb", "kg"}

CAMPOS_IMPORTABLES = (
    "capitulo",
    "partida",
    "descripcion",
    "unidad",
    "cantidad",
    "precio",
    "categoria",
    "tipo_partida",
)

ETIQUETAS_CAMPOS = {
    "capitulo": "Capítulo",
    "partida": "Partida",
    "descripcion": "Descripción",
    "unidad": "Unidad",
    "cantidad": "Cantidad",
    "precio": "Precio unitario",
    "categoria": "Categoría",
    "tipo_partida": "Tipo de partida",
}

# Los sinónimos se normalizan sin tildes, espacios ni puntuación.
ALIAS_CAMPOS = {
    "capitulo": {"capitulo", "cap", "chapter", "grupo", "seccion"},
    "partida": {"partida", "nombre", "item", "concepto", "trabajo", "descripcioncorta"},
    "descripcion": {"descripcion", "detalle", "descripciontecnica", "observaciones"},
    "unidad": {"unidad", "und", "u", "unit"},
    "cantidad": {"cantidad", "cant", "qty", "cantidadtotal"},
    "precio": {"precio", "preciounitario", "punitario", "valorunitario", "unitprice", "preciousd"},
    "categoria": {"categoria", "category", "familia"},
    "tipo_partida": {"tipo", "tipopartida", "tipoitem", "itemtype"},
}

TIPOS_PARTIDA = {
    "included": "included",
    "incluida": "included",
    "incluido": "included",
    "optional": "optional",
    "opcional": "optional",
    "alternative": "alternative",
    "alternativa": "alternative",
    "excluded": "excluded",
    "excluida": "excluded",
    "excluido": "excluded",
    "noincluida": "excluded",
    "noincluido": "excluded",
    "provisional": "provisional",
    "measurement": "measurement",
    "medicion": "measurement",
    "sujetoamedicion": "measurement",
    "sujetaamedicion": "measurement",
}


class ErrorImportacion(ValueError):
    """Error comprensible para mostrar en el asistente de importación."""


def normalizar(valor: Any) -> str:
    texto = unicodedata.normalize("NFD", str(valor or "").strip().lower())
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", texto)


def texto_celda(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def numero_local(valor: Any) -> float | None:
    """Convierte 1.234,50 y 1,234.50 sin adivinar de forma peligrosa."""
    if valor is None or str(valor).strip() == "":
        return None
    texto = str(valor).strip().replace(" ", "").replace("$", "").replace("Bs", "")
    if "," in texto and "." in texto:
        # El último separador corresponde normalmente al decimal.
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        numero = float(texto)
    except (TypeError, ValueError):
        return None
    return numero if numero == numero and abs(numero) != float("inf") else None


def _limitar_matriz(filas: list[list[Any]]) -> list[list[str]]:
    if len(filas) > MAX_FILAS + 1:
        raise ErrorImportacion(f"El archivo supera el límite de {MAX_FILAS} filas importables.")
    resultado = []
    for fila in filas:
        fila = list(fila[:MAX_COLUMNAS])
        limpia = [texto_celda(celda) for celda in fila]
        if any(limpia):
            resultado.append(limpia)
    if not resultado:
        raise ErrorImportacion("No se encontraron filas con datos.")
    return resultado


def leer_csv(contenido: bytes) -> list[list[str]]:
    if len(contenido) > MAX_BYTES:
        raise ErrorImportacion("El archivo CSV supera el límite de 8 MB.")
    texto = ""
    for codificacion in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = contenido.decode(codificacion)
            break
        except UnicodeDecodeError:
            continue
    if not texto:
        raise ErrorImportacion("No se pudo leer el CSV. Guárdalo como UTF-8 o CSV de Excel.")
    muestra = texto[:4096]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
        lector = csv.reader(io.StringIO(texto), dialecto)
    except csv.Error:
        delimitador = "\t" if "\t" in muestra else ";" if ";" in muestra else ","
        lector = csv.reader(io.StringIO(texto), delimiter=delimitador)
    return _limitar_matriz(list(lector))


def leer_xlsx(contenido: bytes) -> list[list[str]]:
    if len(contenido) > MAX_BYTES:
        raise ErrorImportacion("El archivo Excel supera el límite de 8 MB.")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # protección para instalaciones sin dependencia aún
        raise ErrorImportacion(
            "Falta el componente para leer archivos Excel (.xlsx). "
            "Cierra la aplicación y vuelve a abrirla con INICIAR.bat / INICIAR.sh "
            "(instalan las dependencias automáticamente), o ejecuta: pip install openpyxl"
        ) from exc
    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(min_row=1, max_row=MAX_FILAS + 2, max_col=MAX_COLUMNAS, values_only=True))
        libro.close()
    except Exception as exc:
        raise ErrorImportacion("No se pudo abrir el archivo Excel. Usa un .xlsx válido sin contraseña.") from exc
    return _limitar_matriz(filas)


# ---------------------------------------------------------------------------
# Formato de descompuestos CYPE (DPT020.xlsx)
# ---------------------------------------------------------------------------

# Campos de la tabla de recursos de un descompuesto CYPE y los alias con los
# que el exportador nombra cada columna en la fila de encabezados.
CAMPOS_CYPE = ("codigo", "unidad", "descripcion", "rendimiento", "precio", "importe")
ALIAS_COLUMNAS_CYPE = {
    "codigo": {"codigo"},
    "unidad": {"unidad"},
    "descripcion": {"descripcion"},
    "rendimiento": {"rendimiento"},
    "precio": {"preciounitario", "precio"},
    "importe": {"importe"},
}
COSTES_CYPE = ("materiales", "mano_obra", "complementarios", "otros")


def _cype_encabezado(hoja) -> tuple[int, dict[str, int]] | None:
    """Devuelve ``(fila, posiciones)`` de la cabecera CYPE o ``None``.

    CYPE no usa un layout fijo: ``DPT020.xlsx`` ocupa 8 columnas (Unidad en
    D, Rendimiento en F…) mientras que otros descompuestos como ``RBE010``
    ocupan 10 columnas con separadores estrechos (Unidad en C, Rendimiento
    en G…). En vez de posiciones escritas a mano, se busca la fila que
    contiene los seis encabezados característicos y se registra la columna
    exacta de cada campo. Así no se pierde ninguna fila ni columna sea cual
    sea la variante del exportador.
    """
    limite = min(hoja.max_row, MAX_FILAS_CYPE)
    ancho = min(hoja.max_column, MAX_COLUMNAS_CYPE)
    for fila in range(1, limite + 1):
        posiciones: dict[str, int] = {}
        for columna in range(1, ancho + 1):
            nombre = normalizar(hoja.cell(fila, columna).value)
            if not nombre:
                continue
            for campo, alias in ALIAS_COLUMNAS_CYPE.items():
                if campo not in posiciones and nombre in alias:
                    posiciones[campo] = columna - 1
                    break
        if all(campo in posiciones for campo in CAMPOS_CYPE):
            return fila, posiciones
    return None


def es_formato_cype_xlsx(contenido: bytes) -> bool:
    """Comprueba sin alterar datos si el libro tiene alguna hoja CYPE."""
    if len(contenido) > MAX_BYTES:
        return False
    try:
        from openpyxl import load_workbook
        # La detección consulta posiciones concretas de la hoja; se abre en
        # modo normal para que funcione también con libros de CYPE que usan
        # rangos combinados.
        libro = load_workbook(io.BytesIO(contenido), read_only=False, data_only=False)
        try:
            return any(_cype_encabezado(hoja) is not None for hoja in libro.worksheets)
        finally:
            libro.close()
    except Exception:
        return False


def _texto_valor_cype(valor: Any) -> str:
    """Representación estable para cada celda guardada en la base de datos."""
    return texto_celda(valor)


def _categoria_coste_cype(grupo: str, codigo: str = "") -> str:
    """Clasifica un recurso por su grupo (y, como apoyo, por el prefijo CYPE
    del código: ``mo`` mano de obra, ``mt`` materiales). Los grupos no
    reconocidos (maquinaria, medios auxiliares, equipo…) caen en «otros»."""
    texto = normalizar(grupo)
    clave_codigo = normalizar(codigo)
    if "manodeobra" in texto or "personal" in texto or clave_codigo.startswith("mo"):
        return "mano_obra"
    if "material" in texto or clave_codigo.startswith("mt"):
        return "materiales"
    if "complementario" in texto:
        return "complementarios"
    if any(k in texto for k in ["maquin", "equipo", "auxiliar", "medio", "herramient"]) or clave_codigo.startswith(("mq", "maq", "eq")):
        return "otros"
    return "otros"


def categoria_coste_cype(grupo: str, codigo: str = "") -> str:
    """Versión pública de la clasificación de gastos de un recurso CYPE."""
    return _categoria_coste_cype(grupo, codigo)


def _es_numero_entero(valor: str) -> bool:
    try:
        return float(valor).is_integer()
    except (TypeError, ValueError):
        return False


def _celda_cype(celdas: list[str], posiciones: dict[str, int], campo: str) -> str:
    indice = posiciones[campo]
    return celdas[indice] if indice < len(celdas) else ""


def _buscar_cabecera_partida(filas: list[dict], fila_encabezados: int) -> dict | None:
    """Encuentra código, unidad y título de la partida anteriores a la tabla.

    La fila de la partida usa una disposición propia e independiente de las
    columnas de la tabla de recursos: el código va en A y la unidad en B en
    todos los layouts. El título, en cambio, se desplaza según las celdas
    combinadas (C en DPT020, D en RBE010), así que se toma como la primera
    celda con contenido a partir de la columna C.
    """
    for fila in reversed(filas[: max(0, fila_encabezados - 1)]):
        celdas = fila["celdas"]
        codigo = celdas[0] if len(celdas) > 0 else ""
        unidad = celdas[1] if len(celdas) > 1 else ""
        titulo = next((celda for celda in celdas[2:] if celda), "")
        if codigo and unidad and titulo:
            return {"fila": fila["numero"], "codigo": codigo, "unidad": unidad, "titulo": titulo}
    return None


def _clasificar_filas_cype(
    filas: list[dict], fila_encabezados: int, posiciones: dict[str, int]
) -> tuple[list[dict], dict[str, float], float]:
    """Conserva y clasifica cada fila sin eliminar ninguna.

    Solo se agregan campos derivados (tipo, grupo y valores numéricos) a la
    matriz original. Los importes para costes se toman exclusivamente de
    recursos, nunca de los subtotales, para evitar duplicarlos. Las columnas
    se resuelven desde el mapa de posiciones del encabezado: funciona con el
    layout de 8 columnas (DPT020), con el de 10 columnas (RBE010) y con
    cualquier variante que respete los encabezados CYPE.
    """
    grupo_actual = ""
    costes = {categoria: 0.0 for categoria in COSTES_CYPE}
    importe_directo: float | None = None
    ancho_minimo = max(posiciones.values()) + 1

    for fila in filas:
        celdas = fila["celdas"]
        # La anchura del layout se mantiene incluso si no hay valor, para que
        # una fila combinada o vacía sea reversible.
        celdas += [""] * max(0, ancho_minimo - len(celdas))
        codigo = _celda_cype(celdas, posiciones, "codigo")
        unidad = _celda_cype(celdas, posiciones, "unidad")
        descripcion = _celda_cype(celdas, posiciones, "descripcion")
        rendimiento = _celda_cype(celdas, posiciones, "rendimiento")
        precio = _celda_cype(celdas, posiciones, "precio")
        importe = _celda_cype(celdas, posiciones, "importe")
        texto_fila = " ".join(celdas)
        normal_fila = normalizar(texto_fila)
        fila["grupo"] = grupo_actual
        fila["codigo"] = ""
        fila["unidad"] = ""
        fila["descripcion"] = ""
        fila["rendimiento"] = None
        fila["precio_unitario"] = None
        fila["importe"] = None

        if fila["numero"] == fila_encabezados:
            fila["tipo"] = "encabezado"
            continue
        if not any(celdas):
            fila["tipo"] = "vacia"
            continue
        if fila["numero"] < fila_encabezados:
            fila["tipo"] = "cabecera"
            continue

        # Ejemplo: A9=1 y E9='Mano de obra'. Esta comprobación va antes de
        # «costes directos»: el nombre de un grupo puede ser precisamente
        # «Costes directos complementarios». También antes de «recurso»,
        # porque el número de grupo va en la columna de código.
        if _es_numero_entero(codigo) and descripcion and not unidad:
            grupo_actual = descripcion
            fila["grupo"] = grupo_actual
            fila["tipo"] = "grupo"
            fila["codigo"] = codigo
            fila["descripcion"] = descripcion
            continue

        # Los subtotales ocupan las columnas finales y nunca se suman como
        # recursos.
        if "subtotal" in normal_fila:
            fila["tipo"] = "subtotal"
            fila["descripcion"] = descripcion or rendimiento
            fila["importe"] = numero_local(importe)
            continue

        rendimiento_num = numero_local(rendimiento)
        precio_num = numero_local(precio)
        importe_num = numero_local(importe)
        # Los complementarios CYPE se expresan como una línea de porcentaje:
        # no llevan código A, pero sí unidad (%), descripción y tres valores.
        if unidad and descripcion and (rendimiento_num is not None or precio_num is not None or importe_num is not None):
            fila["tipo"] = "recurso"
            fila["grupo"] = grupo_actual
            fila["codigo"] = codigo
            fila["unidad"] = unidad
            fila["descripcion"] = descripcion
            fila["rendimiento"] = rendimiento_num
            fila["precio_unitario"] = precio_num
            # Si Excel no ha recalculado la fórmula, se conserva la fórmula
            # original y se usa el cálculo equivalente para el coste inicial.
            fila["importe"] = importe_num
            if fila["importe"] is None and rendimiento_num is not None and precio_num is not None:
                fila["importe"] = round(rendimiento_num * precio_num, 2)
            if fila["importe"] is not None:
                categoria = _categoria_coste_cype(grupo_actual, codigo)
                costes[categoria] += fila["importe"]
                fila["categoria"] = categoria
            continue

        # Total final del descompuesto, por ejemplo «Costes directos (1+2)».
        if "costesdirectos" in normal_fila or "costosdirectos" in normal_fila:
            fila["tipo"] = "total"
            fila["descripcion"] = descripcion or rendimiento
            fila["importe"] = numero_local(importe)
            if fila["importe"] is not None:
                importe_directo = fila["importe"]
            continue

        fila["tipo"] = "otro"
        fila["descripcion"] = descripcion or texto_fila

    coste_recursos = round(sum(costes.values()), 2)
    if importe_directo is None:
        importe_directo = coste_recursos
    # Los complementarios y redondeos aparecen como filas de porcentaje o
    # totales. Para que el coste interno de la partida coincida exactamente
    # con «Costes directos», la diferencia queda en «otros».
    diferencia = round(importe_directo - coste_recursos, 2)
    if diferencia:
        costes["otros"] = round(costes["otros"] + diferencia, 2)
    costes = {clave: round(max(0.0, valor), 2) for clave, valor in costes.items()}
    return filas, costes, round(importe_directo, 2)


def analizar_cype_xlsx(contenido: bytes) -> dict:
    """Lee descompuestos CYPE preservando la estructura del libro.

    Se abren dos vistas del mismo libro: una con las fórmulas y otra con el
    valor cacheado por Excel. De esta manera se conservan tanto el resultado
    mostrado como la fórmula original sin intentar reescribirla o perderla.
    """
    if len(contenido) > MAX_BYTES:
        raise ErrorImportacion("El archivo Excel supera el límite de 8 MB.")
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ErrorImportacion(
            "Falta el componente para leer archivos Excel (.xlsx). "
            "Cierra la aplicación y vuelve a abrirla con INICIAR.bat / INICIAR.sh "
            "(instalan las dependencias automáticamente), o ejecuta: pip install openpyxl"
        ) from exc

    try:
        libro_formula = load_workbook(io.BytesIO(contenido), read_only=False, data_only=False)
        libro_valores = load_workbook(io.BytesIO(contenido), read_only=False, data_only=True)
    except Exception as exc:
        raise ErrorImportacion("No se pudo abrir el archivo Excel. Usa un .xlsx válido sin contraseña.") from exc

    partidas: list[dict] = []
    try:
        for hoja_formula in libro_formula.worksheets:
            hoja_valores = libro_valores[hoja_formula.title]
            encabezado = _cype_encabezado(hoja_formula)
            if encabezado is None:
                continue
            fila_encabezados, posiciones = encabezado
            if hoja_formula.max_row > MAX_FILAS_CYPE or hoja_formula.max_column > MAX_COLUMNAS_CYPE:
                raise ErrorImportacion(
                    f"La hoja «{hoja_formula.title}» tiene {hoja_formula.max_row} filas y "
                    f"{hoja_formula.max_column} columnas. No se importó parcialmente para no perder datos."
                )

            columnas = [get_column_letter(indice) for indice in range(1, hoja_formula.max_column + 1)]
            filas: list[dict] = []
            for numero_fila in range(1, hoja_formula.max_row + 1):
                celdas: list[str] = []
                formulas: dict[str, str] = {}
                for columna in range(1, hoja_formula.max_column + 1):
                    celda_formula = hoja_formula.cell(numero_fila, columna)
                    celda_valor = hoja_valores.cell(numero_fila, columna)
                    valor = celda_valor.value
                    formula = celda_formula.value if celda_formula.data_type == "f" else None
                    # Cuando el archivo procede de una herramienta que no
                    # guardó la caché, no fingimos un valor: queda vacío y la
                    # fórmula sigue disponible en su columna exacta.
                    celdas.append(_texto_valor_cype(valor))
                    if formula:
                        formulas[get_column_letter(columna)] = str(formula)
                filas.append({"numero": numero_fila, "celdas": celdas, "formulas": formulas})

            cabecera = _buscar_cabecera_partida(filas, fila_encabezados)
            if cabecera is None:
                raise ErrorImportacion(
                    f"Se detectó la tabla CYPE en «{hoja_formula.title}», pero no su cabecera de partida (código, unidad y descripción)."
                )
            filas, costes, coste_directo = _clasificar_filas_cype(filas, fila_encabezados, posiciones)
            descripcion_larga = ""
            for fila in filas:
                if cabecera["fila"] < fila["numero"] < fila_encabezados:
                    for celda in fila["celdas"]:
                        if len(celda) > len(descripcion_larga):
                            descripcion_larga = celda
            # Es preferible no duplicar el título como descripción si la hoja
            # no incluyó una descripción técnica extensa.
            if normalizar(descripcion_larga) == normalizar(cabecera["titulo"]):
                descripcion_larga = ""

            partidas.append({
                "hoja": hoja_formula.title,
                "codigo": cabecera["codigo"],
                "unidad": cabecera["unidad"],
                "nombre": cabecera["titulo"],
                "descripcion": descripcion_larga,
                "fila_cabecera": cabecera["fila"],
                "fila_encabezados": fila_encabezados,
                "columnas": columnas,
                "filas": filas,
                "rangos_combinados": [str(rango) for rango in hoja_formula.merged_cells.ranges],
                "dimension_original": hoja_formula.calculate_dimension(),
                "costes": costes,
                "coste_directo_unitario": coste_directo,
            })
    finally:
        libro_formula.close()
        libro_valores.close()

    if not partidas:
        raise ErrorImportacion("No se encontró una hoja con el formato de descompuesto CYPE (Código, Unidad, Descripción, Rendimiento, Precio unitario e Importe).")
    return {
        "formato": "cype_descompuesto",
        "partidas": partidas,
        "partidas_detectadas": len(partidas),
        "filas_detectadas": sum(len(partida["filas"]) for partida in partidas),
    }


def _redondear2(valor: float) -> float:
    """Equivalente al ROUND(x; 2) de Excel (mitad hacia arriba)."""
    return float(Decimal(str(valor)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def recalcular_descompuesto_cype(filas) -> dict:
    """Recalcula la cascada de costes de un descompuesto CYPE.

    Cada recurso cuesta, por unidad de partida:

        Importe = Rendimiento × Precio unitario

    p. ej. 0,537 h/m² × 24,41 $/h = 13,11 $/m² de mano de obra, u
    8,5 kg/m² × 0,23 $/kg = 1,96 $/m² de material. El Rendimiento se
    refiere a la unidad de la columna «Unidad» de esa fila (h, kg, m³…) y la
    partida da el precio final por su propia unidad (m², m³…). Después:

    - Subtotal de grupo = suma de los importes de sus recursos.
    - Costes directos complementarios: fila de porcentaje cuyo importe es
      % × (suma de los demás subtotales) / 100.
    - Coste directo de la partida = todos los subtotales + complementarios.

    Son las mismas reglas que aplican las fórmulas del Excel original, así
    que al editar un rendimiento o un precio unitario toda la cadena se
    actualiza igual que en la hoja. ``filas`` admite objetos ORM o dicts con
    ``tipo``, ``grupo``, ``codigo``, ``unidad``, ``rendimiento`` y
    ``precio_unitario``; no se modifica nada: se devuelven los valores
    derivados indexados por posición en la lista.
    """
    def campo(fila, nombre):
        return fila.get(nombre) if isinstance(fila, dict) else getattr(fila, nombre, None)

    def es_porcentaje(fila):
        # Ojo: no usar normalizar() aquí, porque convierte «%» en cadena
        # vacía al quitar la puntuación.
        return str(campo(fila, "unidad") or "").strip() == "%"

    def categoria_fila(fila):
        """Categoría de coste de una fila: la explícita si existe (filas
        creadas/editadas a mano), si no la derivada del grupo/código CYPE."""
        categoria = str(campo(fila, "categoria") or "").strip()
        if categoria in COSTES_CYPE:
            return categoria
        return _categoria_coste_cype(campo(fila, "grupo") or "", campo(fila, "codigo") or "")

    importes: dict[int, float] = {}
    subtotal_grupo: dict[str, float] = {}
    costes = {categoria: 0.0 for categoria in COSTES_CYPE}

    # 1) Recursos convencionales: importe = rendimiento × precio unitario.
    for indice, fila in enumerate(filas):
        if campo(fila, "tipo") != "recurso" or es_porcentaje(fila):
            continue
        rendimiento = float(campo(fila, "rendimiento") or 0.0)
        precio = float(campo(fila, "precio_unitario") or 0.0)
        importe = _redondear2(rendimiento * precio)
        importes[indice] = importe
        grupo = campo(fila, "grupo") or ""
        subtotal_grupo[grupo] = _redondear2(subtotal_grupo.get(grupo, 0.0) + importe)
        categoria = categoria_fila(fila)
        costes[categoria] = _redondear2(costes[categoria] + importe)

    # 2) Base de los complementarios: la suma de los subtotales de los demás
    #    grupos (es la columna «Precio unitario» de la fila % en el Excel).
    #    Se apoya en las filas de subtotal cuando existen; si no, en los
    #    totales acumulados por grupo.
    base = 0.0
    hay_subtotales = False
    for fila in filas:
        if campo(fila, "tipo") != "subtotal":
            continue
        hay_subtotales = True
        grupo = campo(fila, "grupo") or ""
        if categoria_fila(fila) != "complementarios":
            base = _redondear2(base + subtotal_grupo.get(grupo, 0.0))
    if not hay_subtotales:
        for grupo, suma in subtotal_grupo.items():
            if _categoria_coste_cype(grupo) != "complementarios":
                base = _redondear2(base + suma)

    # 3) Filas de porcentaje: importe = % × base / 100.
    precios_complementarios: dict[int, float] = {}
    for indice, fila in enumerate(filas):
        if campo(fila, "tipo") != "recurso" or not es_porcentaje(fila):
            continue
        porcentaje = float(campo(fila, "rendimiento") or 0.0)
        precios_complementarios[indice] = base
        importe = _redondear2(porcentaje * base / 100.0)
        importes[indice] = importe
        costes["complementarios"] = _redondear2(costes["complementarios"] + importe)

    # 4) Subtotales y total: valores derivados para mantener la vista y los
    #    registros sincronizados con el cálculo.
    subtotales: dict[int, float] = {}
    for indice, fila in enumerate(filas):
        if campo(fila, "tipo") == "subtotal":
            grupo = campo(fila, "grupo") or ""
            subtotales[indice] = subtotal_grupo.get(grupo, 0.0)

    coste_directo = _redondear2(sum(importes.values()))
    return {
        "importes": importes,
        "subtotales": subtotales,
        "precios_complementarios": precios_complementarios,
        "base_complementarios": base,
        "costes": costes,
        "coste_directo": coste_directo,
    }


def posiciones_columnas_cype(celdas_encabezado) -> dict[str, int]:
    """Deriva el mapa campo → columna desde las celdas de la fila de
    encabezados guardada (misma lógica que la detección en el libro)."""
    posiciones: dict[str, int] = {}
    for indice, valor in enumerate(celdas_encabezado or []):
        nombre = normalizar(valor)
        if not nombre:
            continue
        for campo, alias in ALIAS_COLUMNAS_CYPE.items():
            if campo not in posiciones and nombre in alias:
                posiciones[campo] = indice
                break
    return posiciones if all(campo in posiciones for campo in CAMPOS_CYPE) else {}


def leer_texto(texto: str) -> list[list[str]]:
    return leer_csv((texto or "").encode("utf-8"))


def analizar_matriz(matriz: list[list[str]], tiene_encabezados: bool = True) -> dict:
    matriz = _limitar_matriz(matriz)
    columnas = max(len(fila) for fila in matriz)
    matriz = [fila + [""] * (columnas - len(fila)) for fila in matriz]
    if tiene_encabezados:
        encabezados = [valor or f"Columna {i + 1}" for i, valor in enumerate(matriz[0])]
        filas = matriz[1:]
    else:
        encabezados = [f"Columna {i + 1}" for i in range(columnas)]
        filas = matriz
    if not filas:
        raise ErrorImportacion("Solo se encontró la fila de encabezados; agrega al menos una partida.")
    if len(filas) > MAX_FILAS:
        raise ErrorImportacion(f"La importación supera el límite de {MAX_FILAS} filas.")
    return {
        "encabezados": encabezados,
        "filas": filas[:MAX_FILAS],
        "mapeo_sugerido": detectar_mapeo(encabezados),
    }


def detectar_mapeo(encabezados: list[str]) -> dict[str, int | None]:
    usados: set[int] = set()
    resultado: dict[str, int | None] = {}
    for campo in CAMPOS_IMPORTABLES:
        indice = None
        alias = ALIAS_CAMPOS[campo]
        for i, encabezado in enumerate(encabezados):
            normal = normalizar(encabezado)
            if i not in usados and (normal in alias or any(a in normal for a in alias if len(a) > 3)):
                indice = i
                usados.add(i)
                break
        resultado[campo] = indice
    return resultado


def normalizar_mapeo(mapeo: dict[str, Any], columnas: int) -> dict[str, int | None]:
    limpio: dict[str, int | None] = {}
    usados: set[int] = set()
    for campo in CAMPOS_IMPORTABLES:
        valor = (mapeo or {}).get(campo)
        try:
            indice = int(valor) if valor not in (None, "", "-1") else None
        except (TypeError, ValueError):
            indice = None
        if indice is not None and (indice < 0 or indice >= columnas or indice in usados):
            indice = None
        if indice is not None:
            usados.add(indice)
        limpio[campo] = indice
    return limpio


def _celda(fila: list[str], mapeo: dict[str, int | None], campo: str) -> str:
    indice = mapeo.get(campo)
    return fila[indice].strip() if indice is not None and indice < len(fila) else ""


def _advertencia(fila: int, mensaje: str) -> dict:
    return {"fila": fila, "mensaje": mensaje}


def validar_filas(
    filas: list[list[str]],
    mapeo_original: dict[str, Any],
    capitulos_existentes: list[str] | None = None,
    primera_fila: int = 2,
) -> dict:
    """Valida y normaliza filas antes de que `main` cree modelos SQLAlchemy."""
    columnas = max((len(fila) for fila in filas), default=0)
    mapeo = normalizar_mapeo(mapeo_original, columnas)
    errores: list[dict] = []
    advertencias: list[dict] = []
    normalizadas: list[dict] = []
    if mapeo.get("partida") is None:
        return {
            "mapeo": mapeo,
            "errores": [_advertencia(0, "Asigna una columna para «Partida» antes de continuar.")],
            "advertencias": [],
            "filas": [],
        }

    existentes = {normalizar(nombre) for nombre in (capitulos_existentes or []) if nombre}
    capitulos_advertidos: set[str] = set()
    duplicados: set[tuple] = set()
    for offset, fila in enumerate(filas):
        numero_fila = primera_fila + offset
        fila = [texto_celda(celda) for celda in fila]
        if not any(fila):
            continue
        nombre = _celda(fila, mapeo, "partida")
        if not nombre:
            errores.append(_advertencia(numero_fila, "La partida no puede estar vacía."))
            continue
        capitulo = _celda(fila, mapeo, "capitulo") or "CAPÍTULO GENERAL"
        if not _celda(fila, mapeo, "capitulo"):
            advertencias.append(_advertencia(numero_fila, "No se indicó capítulo; se usará «CAPÍTULO GENERAL»."))
        clave_capitulo = normalizar(capitulo)
        if existentes and clave_capitulo not in existentes and clave_capitulo not in capitulos_advertidos:
            advertencias.append(_advertencia(numero_fila, f"El capítulo «{capitulo}» no existe en el presupuesto destino; se creará."))
            capitulos_advertidos.add(clave_capitulo)

        cantidad_texto = _celda(fila, mapeo, "cantidad")
        cantidad = numero_local(cantidad_texto)
        if cantidad is None:
            cantidad = 1.0
            advertencias.append(_advertencia(numero_fila, "Cantidad vacía o inválida; se usará 1."))
        elif cantidad < 0:
            errores.append(_advertencia(numero_fila, "La cantidad no puede ser negativa."))
            continue

        precio_texto = _celda(fila, mapeo, "precio")
        precio = numero_local(precio_texto)
        if precio is None:
            precio = 0.0
            advertencias.append(_advertencia(numero_fila, "Precio vacío o inválido; se usará 0,00."))
        elif precio < 0:
            errores.append(_advertencia(numero_fila, "El precio no puede ser negativo."))
            continue

        unidad = _celda(fila, mapeo, "unidad") or "ud"
        if normalizar(unidad) not in {normalizar(valor) for valor in UNIDADES_COMUNES}:
            advertencias.append(_advertencia(numero_fila, f"La unidad «{unidad}» no es habitual; se conservará tal como fue importada."))
        tipo_texto = _celda(fila, mapeo, "tipo_partida")
        tipo = TIPOS_PARTIDA.get(normalizar(tipo_texto), "included")
        if tipo_texto and normalizar(tipo_texto) not in TIPOS_PARTIDA:
            advertencias.append(_advertencia(numero_fila, f"Tipo «{tipo_texto}» no reconocido; se importará como incluida."))
        item = {
            "capitulo": capitulo.strip().upper(),
            "nombre": nombre,
            "descripcion": _celda(fila, mapeo, "descripcion"),
            "unidad": unidad,
            "cantidad": cantidad,
            "precio": precio,
            "categoria": _celda(fila, mapeo, "categoria") or "General",
            "tipo_partida": tipo,
        }
        clave = (normalizar(item["capitulo"]), normalizar(nombre), normalizar(unidad), cantidad, precio)
        if clave in duplicados:
            advertencias.append(_advertencia(numero_fila, "Fila duplicada detectada; se conservará como una línea independiente."))
        duplicados.add(clave)
        normalizadas.append(item)

    if not normalizadas and not errores:
        errores.append(_advertencia(0, "No se encontraron partidas importables."))
    return {"mapeo": mapeo, "errores": errores, "advertencias": advertencias, "filas": normalizadas}
